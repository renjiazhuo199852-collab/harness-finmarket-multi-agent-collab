"""Tests for the repeatable synthetic Excel/DB row generators."""

from __future__ import annotations

from datetime import datetime, timezone

from openpyxl import load_workbook

from agent.scripts.generate_fx_synthetic_excel import build_workbook
from agent.scripts.seed_fx_synthetic_data import (
    _bar_rows,
    _macro_rows,
    _news_rows,
    _price_rows,
)


def test_default_shape_has_4h_warmup_and_legacy_headers(tmp_path) -> None:
    path = build_workbook(
        tmp_path / "fx.xlsx",
        datetime(2026, 8, 25, tzinfo=timezone.utc),
    )
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == [
            "latest_prices",
            "market_bars",
            "macro_observations",
            "news_articles",
        ]
        bars = list(workbook["market_bars"].values)
        assert bars[0][:2] == ("bar_time", "frequency")
        assert sum(row[1] == "daily" for row in bars[1:]) == 8 * 400
        assert sum(row[1] == "hourly" for row in bars[1:]) == 8 * 1200
    finally:
        workbook.close()


def test_database_rows_use_registered_columns_and_frequencies() -> None:
    as_of = datetime(2026, 8, 25, tzinfo=timezone.utc)
    assert len(_price_rows(as_of)[0]) == 7
    bars = _bar_rows(as_of=as_of, daily_bars=2, hourly_bars=4)
    assert len(bars[0]) == 10
    assert {row[2] for row in bars} == {"daily", "hourly"}
    assert len(_macro_rows(as_of, 2)[0]) == 12
    assert len(_news_rows(as_of, 2)[0]) == 10
