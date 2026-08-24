"""Excel adapter and deterministic Evidence Factory V2 tests."""

from __future__ import annotations

from datetime import datetime, timedelta, timezone

from openpyxl import Workbook

from src.fx_debate.context import build_evidence_context
from src.fx_debate.evidence_factory import FxEvidenceFactory, _complete_four_hour_bars
from src.fx_debate.evidence_sources import (
    ExcelFxEvidenceSource,
    RawFxSnapshot,
    ReaderFxEvidenceSource,
)
from src.fx_debate.models import ResolvedFxDebateRequest, RunOptions
from src.fx_debate.store import FxEvidenceStore


def _context(as_of: datetime):
    request = ResolvedFxDebateRequest(
        status="resolved",
        asset_class="fx",
        instrument_type="spot",
        pair_class="major",
        canonical_symbol="EURUSD",
        display_symbol="EUR/USD",
        base_currency="EUR",
        quote_currency="USD",
        requested_base_currency="EUR",
        requested_quote_currency="USD",
        inverted=False,
        horizon="2 weeks",
        timeframe="4H/1D",
    )
    return build_evidence_context(
        request,
        RunOptions(request_id="req-factory-v2", as_of=as_of),
    )


class _StaticSource:
    def __init__(self, snapshot: RawFxSnapshot) -> None:
        self.snapshot = snapshot

    def load(self, context):
        del context
        return self.snapshot


def test_macro_evidence_ids_include_country_for_same_metric_and_release(
    tmp_path,
) -> None:
    context = _context(datetime(2026, 8, 2, 12, tzinfo=timezone.utc))
    release_time = context.as_of - timedelta(days=1)
    rows = []
    for country, value in (("EU", 2.1), ("US", 2.4)):
        rows.append(
            {
                "metric_id": "CPI",
                "release_time": release_time,
                "frequency": "monthly",
                "value": value,
                "previous_value": value - 0.1,
                "forecast_value": value - 0.05,
                "revised_value": None,
                "unit": "percent",
                "source": "SYNTHETIC",
                "source_identifier": f"{country}-CPI",
                "country": country,
            }
        )

    bundle = FxEvidenceFactory().build(
        context,
        _StaticSource(RawFxSnapshot(source_name="test", macro=rows)),
    )
    macro_items = [item for item in bundle.evidence if item.domain == "macro"]

    assert len(macro_items) == 2
    assert len({item.evidence_id for item in macro_items}) == 2
    FxEvidenceStore(tmp_path, context.evidence_context_id).register(bundle.evidence)


def test_degraded_bundle_builds_presentable_background_without_four_hour_data() -> None:
    as_of = datetime(2026, 8, 22, 12, tzinfo=timezone.utc)
    context = _context(as_of)
    release_time = as_of - timedelta(days=23)
    rows = [
        {
            "metric_id": f"{country}_{metric_id}",
            "release_time": release_time,
            "frequency": "monthly",
            "value": value,
            "previous_value": None,
            "forecast_value": None,
            "revised_value": None,
            "unit": "percent" if "UNEMPLOYMENT" in metric_id or "CPI" in metric_id else "index",
            "source": "EXCEL",
            "source_identifier": f"{country}-{metric_id}",
            "country": country,
        }
        for metric_id, country, value in (
            ("PMI_MANUFACTURING", "EU", 52.0),
            ("PMI_MANUFACTURING", "US", 53.3),
            ("PMI_SERVICES", "EU", 51.6),
            ("PMI_SERVICES", "US", 54.0),
            ("UNEMPLOYMENT", "EU", 6.27),
            ("UNEMPLOYMENT", "US", 4.27),
            ("CPI_YOY", "EU", 3.0),
            ("CPI_YOY", "US", 3.86),
        )
    ]

    bundle = FxEvidenceFactory().build(
        context,
        _StaticSource(
            RawFxSnapshot(
                source_name="excel",
                macro=rows,
                bars=[
                    {
                        "bar_time": as_of - timedelta(days=offset),
                        "frequency": "daily",
                        "open": 1.10,
                        "high": 1.11,
                        "low": 1.09,
                        "close": 1.10,
                    }
                    for offset in range(1, 22)
                ],
            )
        ),
    )

    assert bundle.presentation.market_background == "美元历史基本面背景偏强，EUR/USD 宏观背景偏空"
    assert bundle.presentation.background_strength == "low"
    assert "无法确认：4H 无数据，1D 仅 21 根" in bundle.presentation.technical_confirmation
    assert "完整确认仍需 50 根" in bundle.presentation.technical_confirmation
    assert bundle.presentation.data_quality == "degraded"
    assert any("US PMI 高于 EU PMI" in item for item in bundle.presentation.usable_evidence)
    assert any("1D 观察事实" in item for item in bundle.presentation.usable_evidence)
    assert any("forecast 缺失" in item for item in bundle.presentation.limitations)


def _write_workbook(path, as_of: datetime, *, complete: bool) -> None:
    excel_as_of = as_of.replace(tzinfo=None)
    workbook = Workbook()
    workbook.remove(workbook.active)

    prices = workbook.create_sheet("latest_prices")
    prices.append(
        [
            "id",
            "price_time",
            "last",
            "bid",
            "ask",
            "mid",
            "source",
            "source_identifier",
            "updated_at",
        ]
    )
    prices.append(
        [
            1,
            excel_as_of - timedelta(minutes=5),
            1.1,
            1.0999,
            1.1001,
            1.1,
            "TEST",
            "EUR=",
            excel_as_of,
        ]
    )
    prices.append(
        [
            2,
            excel_as_of + timedelta(minutes=1),
            9.9,
            9.8,
            10.0,
            9.9,
            "TEST",
            "EUR=",
            excel_as_of,
        ]
    )

    macro = workbook.create_sheet("macro_observations")
    macro.append(
        [
            "id",
            "metric_id",
            "instrument_id",
            "release_time",
            "frequency",
            "value",
            "previous_value",
            "forecast_value",
            "revised_value",
            "source",
            "source_identifier",
            "country",
            "unit",
            "created_at",
            "updated_at",
        ]
    )
    forecast_eu = 50.0 if complete else None
    forecast_us = 51.0 if complete else None
    macro.append(
        [
            1,
            "EU_POLICY_RATE",
            None,
            excel_as_of - timedelta(days=2),
            "daily",
            3.0,
            None,
            None,
            None,
            "TEST",
            "EU_RATE",
            "EU",
            "%",
            excel_as_of,
            excel_as_of,
        ]
    )
    macro.append(
        [
            2,
            "US_POLICY_RATE",
            None,
            excel_as_of - timedelta(days=2),
            "daily",
            4.0,
            None,
            None,
            None,
            "TEST",
            "US_RATE",
            "US",
            "%",
            excel_as_of,
            excel_as_of,
        ]
    )
    macro.append(
        [
            3,
            "EU_PMI_MANUFACTURING",
            None,
            excel_as_of - timedelta(days=1),
            "monthly",
            51.0,
            50.0,
            forecast_eu,
            None,
            "TEST",
            "EU_PMI",
            "EU",
            "index",
            excel_as_of,
            excel_as_of,
        ]
    )
    macro.append(
        [
            4,
            "US_PMI_MANUFACTURING",
            None,
            excel_as_of - timedelta(days=1),
            "monthly",
            50.5,
            50.0,
            forecast_us,
            None,
            "TEST",
            "US_PMI",
            "US",
            "index",
            excel_as_of,
            excel_as_of,
        ]
    )

    bars = workbook.create_sheet("market_bars")
    bars.append(
        [
            "id",
            "date",
            "bar_time",
            "frequency",
            "open",
            "high",
            "low",
            "close",
            "volume",
            "source",
            "source_identifier",
            "updated_at",
        ]
    )
    daily_count = 60 if complete else 22
    for index in range(daily_count):
        timestamp = excel_as_of - timedelta(days=daily_count - index)
        close = 1.05 + index / 10_000
        bars.append(
            [
                index,
                timestamp.date(),
                timestamp,
                "daily",
                close - 0.0002,
                close + 0.0005,
                close - 0.0005,
                close,
                None,
                "TEST",
                "EUR=",
                excel_as_of,
            ]
        )
    if complete:
        start = excel_as_of - timedelta(hours=240)
        for index in range(240):
            timestamp = start + timedelta(hours=index)
            close = 1.08 + index / 100_000
            bars.append(
                [
                    1000 + index,
                    timestamp.date(),
                    timestamp,
                    "hourly",
                    close - 0.0001,
                    close + 0.0003,
                    close - 0.0003,
                    close,
                    None,
                    "TEST",
                    "EUR=",
                    excel_as_of,
                ]
            )

    news = workbook.create_sheet("news_articles")
    news.append(
        [
            "id",
            "article_id",
            "source",
            "publish_time",
            "title",
            "language",
            "sentiment_score",
            "related_entities",
            "keywords",
            "updated_at",
            "content",
            "summary",
        ]
    )
    news.append(
        [
            1,
            "A1",
            "TEST",
            excel_as_of - timedelta(hours=4),
            "ECB keeps guidance unchanged",
            "en",
            None,
            '{"query_tag":"EU"}',
            '{"topic":["EU"]}',
            excel_as_of,
            "body",
            None,
        ]
    )
    news.append(
        [
            2,
            "A2",
            "TEST",
            excel_as_of - timedelta(hours=3),
            "UPDATE 1 ECB keeps guidance unchanged",
            "en",
            None,
            '{"query_tag":"EU"}',
            '{"topic":["EU"]}',
            excel_as_of,
            "body",
            None,
        ]
    )
    workbook.save(path)


def test_export_like_excel_degrades_without_4h_or_macro_forecasts(tmp_path) -> None:
    as_of = datetime(2026, 8, 2, 12, tzinfo=timezone.utc)
    path = tmp_path / "export-like.xlsx"
    _write_workbook(path, as_of, complete=False)

    bundle = FxEvidenceFactory().build(_context(as_of), ExcelFxEvidenceSource(path))

    assert bundle.source_name == "excel"
    assert bundle.manifest.overall_status == "partial"
    assert bundle.technical_regime.timeframes["1D"].state == "indeterminate"
    assert bundle.technical_regime.timeframes["1D"].bar_count == 22
    assert bundle.technical_regime.timeframes["1D"].metrics["latest_close"] == 1.0521
    assert "ema_50" not in bundle.technical_regime.timeframes["1D"].metrics
    assert "observation metrics are available" in (
        bundle.technical_regime.timeframes["1D"].reason or ""
    )
    assert bundle.technical_regime.timeframes["4H"].state == "indeterminate"
    assert bundle.relative_macro_scorecard.status == "partial"
    assert all(
        signal.relationship == "unknown"
        for signal in bundle.relative_macro_scorecard.signals
        if signal.dimension == "growth"
    )
    assert len(bundle.story_clusters) == 1
    assert bundle.story_clusters[0].article_ids == ["A1", "A2"]
    quote = next(item for item in bundle.evidence if item.name == "spot_quote")
    assert quote.value["last"] == 1.1


def test_complete_synthetic_excel_builds_both_technical_timeframes(tmp_path) -> None:
    as_of = datetime(2026, 8, 2, 12, tzinfo=timezone.utc)
    path = tmp_path / "complete.xlsx"
    _write_workbook(path, as_of, complete=True)

    bundle = FxEvidenceFactory().build(_context(as_of), ExcelFxEvidenceSource(path))

    assert bundle.manifest.market.status == "complete"
    assert bundle.technical_regime.timeframes["1D"].bar_count == 60
    assert bundle.technical_regime.timeframes["4H"].bar_count == 60
    assert bundle.technical_regime.timeframes["1D"].state != "indeterminate"
    assert bundle.technical_regime.timeframes["4H"].state != "indeterminate"
    growth = next(
        signal
        for signal in bundle.relative_macro_scorecard.signals
        if signal.dimension == "growth"
    )
    assert growth.relationship == "base_supported"


def test_abnormal_quote_is_exposed_to_the_agent_quality_gate() -> None:
    as_of = datetime(2026, 8, 2, 12, tzinfo=timezone.utc)

    class _Source:
        def load(self, context):
            return RawFxSnapshot(
                source_name="test",
                prices=[
                    {
                        "price_time": context.as_of,
                        "last": 0,
                        "bid": 1.2,
                        "ask": 1.1,
                        "mid": 1.15,
                    }
                ],
            )

    bundle = FxEvidenceFactory().build(_context(as_of), _Source())

    assert bundle.technical_regime.quote_quality == "abnormal"
    assert bundle.manifest.quote.status == "partial"
    assert bundle.evidence[0].quality_status == "abnormal"


def test_four_hour_aggregation_rejects_missing_hour_and_open_bucket() -> None:
    as_of = datetime(2026, 8, 2, 12, tzinfo=timezone.utc)

    def row(hour: int) -> dict[str, object]:
        return {
            "bar_time": datetime(2026, 8, 2, hour, tzinfo=timezone.utc),
            "open": 1.1,
            "high": 1.2,
            "low": 1.0,
            "close": 1.1,
            "volume": 1,
        }

    assert _complete_four_hour_bars([row(0), row(1), row(2), row(4)], as_of) == []
    assert len(_complete_four_hour_bars([row(0), row(1), row(2), row(3)], as_of)) == 1


def test_stale_quote_is_not_presented_as_fresh() -> None:
    as_of = datetime(2026, 8, 2, 12, tzinfo=timezone.utc)

    class _Source:
        def load(self, context):
            return RawFxSnapshot(
                source_name="test",
                prices=[
                    {
                        "price_time": context.as_of - timedelta(days=2),
                        "last": 1.1,
                        "bid": 1.09,
                        "ask": 1.11,
                        "mid": 1.1,
                    }
                ],
            )

    bundle = FxEvidenceFactory().build(_context(as_of), _Source())

    assert bundle.technical_regime.quote_quality == "stale"
    assert bundle.evidence[0].quality_status == "stale"
    assert bundle.raw_preview["market"][0]["quality_status"] == "stale"


def test_reader_adapter_filters_same_day_rows_after_as_of() -> None:
    as_of = datetime(2026, 8, 2, 12, tzinfo=timezone.utc)

    class _Reader:
        @staticmethod
        def get_latest_prices(**kwargs):
            return {
                "prices": [
                    {"price_time": as_of, "last_price": 1.1, "mid_price": 1.1},
                    {
                        "price_time": as_of + timedelta(minutes=1),
                        "last_price": 9.9,
                        "mid_price": 9.9,
                    },
                ]
            }

        @staticmethod
        def get_market_bars(**kwargs):
            return {
                "bars": [
                    {"bar_time": as_of, "frequency": kwargs["frequency"]},
                    {
                        "bar_time": as_of + timedelta(minutes=1),
                        "frequency": kwargs["frequency"],
                    },
                ]
            }

        @staticmethod
        def get_macro_observations(**kwargs):
            return {
                "observations": [
                    {"release_time": as_of},
                    {"release_time": as_of + timedelta(minutes=1)},
                ]
            }

        @staticmethod
        def get_news(**kwargs):
            return {
                "articles": [
                    {"publish_time": as_of},
                    {"publish_time": as_of + timedelta(minutes=1)},
                ]
            }

    snapshot = ReaderFxEvidenceSource(_Reader()).load(_context(as_of))

    assert len(snapshot.prices) == 1
    assert len(snapshot.bars) == 2  # one retained row for each requested frequency
    assert len(snapshot.macro) == 1
    assert len(snapshot.news) == 1
    assert all(row["price_time"] <= as_of for row in snapshot.prices)
    assert all(row["bar_time"] <= as_of for row in snapshot.bars)
