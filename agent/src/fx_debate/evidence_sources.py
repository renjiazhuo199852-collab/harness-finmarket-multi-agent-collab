"""Data-source adapters for building one FX Debate evidence snapshot."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Callable, Protocol

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

from src.fx_debate.analytics import as_utc_datetime
from src.fx_debate.data_query_agent import (
    AiSearchClient,
    DataSearchClient,
    FxDataQueryAgent,
    McpAiSearchClient,
)
from src.fx_debate.models import EvidenceContext
from src.market_data_reader import MarketDataReader


class RawFxSnapshot(BaseModel):
    """Normalized raw rows consumed by :class:`FxEvidenceFactory`."""

    model_config = ConfigDict(extra="forbid", arbitrary_types_allowed=True)

    source_name: str
    prices: list[dict[str, Any]] = Field(default_factory=list)
    bars: list[dict[str, Any]] = Field(default_factory=list)
    macro: list[dict[str, Any]] = Field(default_factory=list)
    news: list[dict[str, Any]] = Field(default_factory=list)
    missing_columns: dict[str, list[str]] = Field(default_factory=dict)
    warnings: list[str] = Field(default_factory=list)


class FxEvidenceSource(Protocol):
    """Small seam between raw market-data adapters and the evidence factory."""

    def load(self, context: EvidenceContext) -> RawFxSnapshot:
        """Load and normalize all raw rows needed by one Debate run."""


_SHEETS = {
    "prices": (
        "latest_prices",
        ("price_time", "last", "bid", "ask", "mid", "source", "source_identifier"),
    ),
    "bars": (
        "market_bars",
        (
            "bar_time",
            "frequency",
            "open",
            "high",
            "low",
            "close",
            "source",
            "source_identifier",
        ),
    ),
    "macro": (
        "macro_observations",
        ("metric_id", "release_time", "frequency", "value", "source", "country"),
    ),
    "news": (
        "news_articles",
        ("article_id", "publish_time", "title", "source", "related_entities"),
    ),
}


class ExcelFxEvidenceSource:
    """Read the database-team export without depending on its local SQL schema."""

    def __init__(self, workbook_path: str | Path) -> None:
        self.workbook_path = Path(workbook_path).expanduser().resolve()

    def load(self, context: EvidenceContext) -> RawFxSnapshot:
        if not self.workbook_path.is_file():
            raise ValueError(
                f"FX Debate Excel file does not exist: {self.workbook_path}"
            )

        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        missing_columns: dict[str, list[str]] = {}
        warnings: list[str] = []
        tables: dict[str, list[dict[str, Any]]] = {}
        try:
            for domain, (sheet_name, required) in _SHEETS.items():
                if sheet_name not in workbook.sheetnames:
                    tables[domain] = []
                    missing_columns[domain] = list(required)
                    warnings.append(f"missing sheet: {sheet_name}")
                    continue
                rows, missing = _read_sheet(workbook[sheet_name], required)
                tables[domain] = rows
                if missing:
                    missing_columns[domain] = missing
        finally:
            workbook.close()

        market_identifiers = _market_identifiers(context)
        macro_countries = _macro_countries(context)
        prices = [
            row
            for row in tables["prices"]
            if str(row.get("source_identifier") or "").upper() in market_identifiers
            and _not_after(row.get("price_time"), context.as_of)
        ]
        bars = [
            row
            for row in tables["bars"]
            if str(row.get("source_identifier") or "").upper() in market_identifiers
            and _not_after(row.get("bar_time"), context.as_of)
        ]
        macro = [
            row
            for row in tables["macro"]
            if str(row.get("country") or "").upper() in macro_countries
            and _not_after(row.get("release_time"), context.as_of)
        ]
        news = [
            row
            for row in tables["news"]
            if _news_matches_pair(row, context)
            and _not_after(row.get("publish_time"), context.as_of)
        ]
        if not prices:
            warnings.append(
                f"no quote rows matched {context.display_symbol} identifiers: "
                f"{', '.join(sorted(market_identifiers))}"
            )
        if not bars:
            warnings.append(
                f"no market bars matched {context.display_symbol} identifiers: "
                f"{', '.join(sorted(market_identifiers))}"
            )
        if not macro:
            warnings.append(
                f"no macro rows matched {context.display_symbol} countries: "
                f"{', '.join(sorted(macro_countries))}"
            )
        return RawFxSnapshot(
            source_name="excel",
            prices=prices,
            bars=bars,
            macro=macro,
            news=news,
            missing_columns=missing_columns,
            warnings=warnings,
        )


class ReaderFxEvidenceSource:
    """Adapt the existing MarketDataReader result shapes to RawFxSnapshot."""

    def __init__(self, reader: MarketDataReader | Any | None = None) -> None:
        self.reader = reader or MarketDataReader()

    def load(self, context: EvidenceContext) -> RawFxSnapshot:
        provider = context.provider_priority[0]
        common = {
            "symbol": context.canonical_symbol,
            "source": provider,
            "start_date": context.market_start_time.date().isoformat(),
            "end_date": context.as_of.date().isoformat(),
        }
        prices = self.reader.get_latest_prices(
            symbol=context.canonical_symbol,
            source=provider,
        ).get("prices", [])
        bars: list[dict[str, Any]] = []
        for frequency, limit in (("daily", 400), ("hourly", 1000)):
            payload = self.reader.get_market_bars(
                **common,
                frequency=frequency,
                limit=limit,
            )
            bars.extend(payload.get("bars", []))
        macro = self.reader.get_macro_observations(
            **common,
            limit=context.macro_observation_limit,
        ).get("observations", [])
        news = self.reader.get_news(
            symbol=context.canonical_symbol,
            source=provider,
            start_date=context.news_start_time.date().isoformat(),
            end_date=context.as_of.date().isoformat(),
            limit=context.news_limit,
        ).get("articles", [])
        normalized_prices = [_normalize_reader_price(row) for row in prices]
        normalized_bars = [_normalize_reader_bar(row) for row in bars]
        return RawFxSnapshot(
            source_name="database",
            prices=[
                row
                for row in normalized_prices
                if _not_after(row.get("price_time"), context.as_of)
            ],
            bars=[
                row
                for row in normalized_bars
                if _not_after(row.get("bar_time"), context.as_of)
            ],
            macro=[
                dict(row)
                for row in macro
                if _not_after(row.get("release_time"), context.as_of)
            ],
            news=[
                dict(row)
                for row in news
                if _not_after(row.get("publish_time"), context.as_of)
            ],
        )


class AiSearchFxEvidenceSource:
    """Use the independent AI Search MCP service as the FX evidence provider.

    The production provider is accessed over local MCP stdio. The optional HTTP
    client remains accepted for compatibility tests and older callers.
    """

    def __init__(
        self,
        client: DataSearchClient | None = None,
        *,
        service_url: str | None = None,
        timeout_seconds: float = 30.0,
        max_rows: int = 250,
        trace_callback: Callable[[dict[str, Any]], None] | None = None,
        mcp_command: str = "",
        mcp_args: str = "",
        mcp_server_module: str = "backend.mcp_server",
        mcp_working_directory: str = "",
        mcp_timeout_seconds: float = 30.0,
    ) -> None:
        if client is None:
            if service_url:
                # 兼容旧调用方；生产路径由 _configured_evidence_source
                # 传入 MCP 配置，不再依赖 HTTP 服务地址。
                client = AiSearchClient(
                    service_url,
                    timeout_seconds=timeout_seconds,
                    max_rows=max_rows,
                    trace_callback=trace_callback,
                )
            else:
                client = McpAiSearchClient.from_repository(
                    command=mcp_command,
                    args_json=mcp_args,
                    server_module=mcp_server_module,
                    working_directory=mcp_working_directory,
                    timeout_seconds=mcp_timeout_seconds or timeout_seconds,
                    max_rows=max_rows,
                    trace_callback=trace_callback,
                )
        self.client = client
        self.agent = FxDataQueryAgent(client)

    def load(self, context: EvidenceContext) -> RawFxSnapshot:
        responses = self.agent.retrieve_for_debate(context)
        warnings: list[str] = []
        tables: dict[str, list[dict[str, Any]]] = {}
        provider = context.provider_priority[0] if context.provider_priority else None
        for domain in ("prices", "bars", "macro", "news"):
            response = responses.get(domain) or {}
            if response.get("status") != "success":
                message = response.get("message") or response.get("code") or "查询未返回数据"
                warnings.append(f"{domain} data service query: {message}")
                tables[domain] = []
                continue
            data = response.get("data")
            if not isinstance(data, list):
                warnings.append(f"{domain} data service response data is not a list")
                tables[domain] = []
                continue
            meta = response.get("meta") if isinstance(response.get("meta"), dict) else {}
            if domain == "prices":
                tables[domain] = [
                    {
                        **row,
                        "source": row.get("source") or meta.get("provider") or provider,
                        "source_identifier": row.get("source_identifier")
                        or meta.get("identifier")
                        or context.canonical_symbol,
                    }
                    for row in data
                    if isinstance(row, dict)
                ]
            elif domain == "bars":
                tables[domain] = [
                    {
                        **row,
                        "bar_time": row.get("bar_time") or row.get("date"),
                        "frequency": row.get("frequency") or meta.get("frequency") or "daily",
                        "source": row.get("source") or meta.get("provider") or provider,
                        "source_identifier": row.get("source_identifier")
                        or meta.get("identifier")
                        or context.canonical_symbol,
                    }
                    for row in data
                    if isinstance(row, dict)
                ]
            elif domain == "macro":
                tables[domain] = [
                    _merge_provider_row(
                        row,
                        metadata_defaults={
                            "release_time": context.as_of,
                            "source": meta.get("provider") or provider,
                            "source_identifier": meta.get("identifier") or context.canonical_symbol,
                        },
                    )
                    for row in data
                    if isinstance(row, dict)
                ]
            else:
                tables[domain] = [
                    _merge_provider_row(
                        row,
                        metadata_defaults={
                            "publish_time": context.as_of,
                            "source": meta.get("provider") or provider,
                            "article_id": "",
                        },
                    )
                    for row in data
                    if isinstance(row, dict)
                ]
        if not tables.get("bars"):
            warnings.append("AI Search market_bars 未返回日线或小时原始 K 线；4H 证据可能不足")
        return RawFxSnapshot(
            source_name="ai_search",
            prices=tables.get("prices", []),
            bars=tables.get("bars", []),
            macro=tables.get("macro", []),
            news=tables.get("news", []),
            warnings=warnings,
        )


def _merge_provider_row(
    row: dict[str, Any],
    *,
    metadata_defaults: dict[str, Any],
) -> dict[str, Any]:
    """Flatten ``{data, metadata}`` rows without losing provenance."""
    data = row.get("data") if isinstance(row.get("data"), dict) else {}
    metadata = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
    # MCP evidence.v1 uses nested data/metadata. Preserve flat rows too so a
    # cached response from the older public contract is not silently reduced
    # to only synthetic defaults.
    flat = row if not data and not metadata else {}
    merged = {**metadata_defaults, **flat, **metadata, **data}
    # Older MCP/public responses exposed the related-macro join role but
    # dropped the country metadata. Keep those responses readable while the
    # evidence contract supplies the canonical country column directly.
    if not merged.get("country"):
        role = str(merged.get("relationship_role") or "").strip().lower()
        if role == "base_currency":
            merged["country"] = "EU"
        elif role == "quote_currency":
            merged["country"] = "US"
    return merged


def _read_sheet(
    sheet: Any, required: tuple[str, ...]
) -> tuple[list[dict[str, Any]], list[str]]:
    header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [
        str(value).strip() if value is not None else "" for value in header_values
    ]
    missing = sorted(set(required) - set(headers))
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values[: len(headers)]):
            break
        rows.append({header: value for header, value in zip(headers, values) if header})
    return rows, missing


def _not_after(value: Any, as_of: Any) -> bool:
    if value is None:
        return False
    try:
        return as_utc_datetime(value) <= as_of
    except (TypeError, ValueError):
        return False


_CURRENCY_COUNTRY = {
    "EUR": "EU",
    "GBP": "UK",
    "JPY": "JP",
    "AUD": "AU",
    "CAD": "CA",
    "CHF": "CH",
    "NZD": "NZ",
    "CNY": "CN",
    "HKD": "HK",
    "SGD": "SG",
    "SEK": "SE",
    "NOK": "NO",
    "MXN": "MX",
    "ZAR": "ZA",
    "USD": "US",
}


def _market_identifiers(context: EvidenceContext) -> set[str]:
    """Return common export identifiers for the requested currency pair.

    The database export stores LSEG-style outright legs (``EUR=``, ``JPY=``)
    rather than a single six-letter symbol.  For a USD cross the non-USD leg
    is therefore the stable lookup key.  Supporting the canonical symbol as a
    fallback also makes synthetic and future Reader exports easy to adapt.
    """
    base = context.base_currency.upper()
    quote = context.quote_currency.upper()
    identifiers = {context.canonical_symbol.upper(), f"{base}{quote}"}
    if base != "USD" and quote != "USD":
        # A non-USD cross must be represented by an outright cross in the
        # export; never combine the two currency legs into a false pair.
        identifiers.update({f"{base}{quote}=R", f"{base}{quote}="})
    else:
        if base != "USD":
            identifiers.add(f"{base}=")
        if quote != "USD":
            identifiers.add(f"{quote}=")
    return identifiers


def _macro_countries(context: EvidenceContext) -> set[str]:
    return {
        _CURRENCY_COUNTRY.get(
            context.base_currency.upper(), context.base_currency.upper()
        ),
        _CURRENCY_COUNTRY.get(
            context.quote_currency.upper(), context.quote_currency.upper()
        ),
    }


def _news_matches_pair(row: dict[str, Any], context: EvidenceContext) -> bool:
    raw = row.get("related_entities")
    if isinstance(raw, dict):
        payload = raw
    elif isinstance(raw, str):
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            payload = {"query_tag": raw}
    else:
        payload = {}
    tag = str(payload.get("query_tag") or "").upper()
    countries = _macro_countries(context)
    symbol = context.canonical_symbol.upper()
    pair_value = str(payload.get("pair") or payload.get("symbol") or "").upper()
    pair_value = pair_value.replace("/", "").replace("-", "")
    if pair_value:
        return pair_value == symbol
    return tag in countries | {"FX", "RISK", symbol, context.display_symbol.upper()}


def _normalize_reader_price(row: dict[str, Any]) -> dict[str, Any]:
    return {
        **row,
        "last": row.get("last_price"),
        "mid": row.get("mid_price"),
    }


def _normalize_reader_bar(row: dict[str, Any]) -> dict[str, Any]:
    return {
        **row,
        "bar_time": row.get("bar_time") or row.get("bar_date"),
    }
