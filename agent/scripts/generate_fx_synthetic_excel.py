"""Generate a complete, deterministic multi-pair workbook for local Debate tests.

The workbook follows the four-table export contract used by the Excel adapter.
It is deliberately synthetic and must never be treated as live market data.
"""

from __future__ import annotations

import argparse
import json
import math
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook

PAIR_SPECS = (
    # canonical symbol, LSEG-style leg identifier, spot, daily drift, macro base
    ("EURUSD", "EUR=", 1.1050, 0.00082, "EU"),
    ("GBPUSD", "GBP=", 1.2850, 0.00070, "UK"),
    ("USDJPY", "JPY=", 154.20, 0.055, "US"),
    ("AUDUSD", "AUD=", 0.6650, 0.00042, "AU"),
    ("USDCAD", "CAD=", 1.3650, -0.00034, "US"),
    ("USDCHF", "CHF=", 0.8920, -0.00028, "US"),
    ("NZDUSD", "NZD=", 0.6120, 0.00035, "NZ"),
    ("EURJPY", "EURJPY=R", 171.50, 0.080, "EU"),
)

DEFAULT_AS_OF = "2026-08-25T00:00:00Z"
DEFAULT_PRICE_DAYS = 30
DEFAULT_DAILY_BARS = 400
DEFAULT_HOURLY_BARS = 1200
DEFAULT_MACRO_VINTAGES = 8
DEFAULT_NEWS_PER_PAIR = 24

COUNTRY_PARAMS = {
    "EU": (2.75, 0.004, 6.2, 2.3),
    "US": (3.50, 0.006, 4.0, 2.5),
    "UK": (4.25, 0.003, 4.4, 2.6),
    "JP": (0.10, 0.002, 2.5, 2.8),
    "AU": (4.35, 0.003, 4.1, 3.2),
    "CA": (3.75, 0.003, 5.8, 2.7),
    "CH": (1.50, 0.002, 2.4, 1.4),
    "NZ": (4.75, 0.003, 5.0, 3.0),
}


def build_workbook(
    output: Path,
    as_of: datetime,
    *,
    price_days: int = DEFAULT_PRICE_DAYS,
    daily_bars: int = DEFAULT_DAILY_BARS,
    hourly_bars: int = DEFAULT_HOURLY_BARS,
    macro_vintages: int = DEFAULT_MACRO_VINTAGES,
    news_per_pair: int = DEFAULT_NEWS_PER_PAIR,
) -> Path:
    """Build the legacy four-sheet workbook with enough warm-up history.

    The knobs are intentionally explicit so tests can build a small workbook while
    the normal command creates enough daily/hourly samples for the FX evidence
    factory (including 4H aggregation). All rows remain synthetic test data.
    """

    for name, value in (
        ("price_days", price_days),
        ("daily_bars", daily_bars),
        ("hourly_bars", hourly_bars),
        ("macro_vintages", macro_vintages),
        ("news_per_pair", news_per_pair),
    ):
        if value < 1:
            raise ValueError(f"{name} must be positive")

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    prices = workbook.create_sheet("latest_prices")
    prices.append(
        ["price_time", "last", "bid", "ask", "mid", "source", "source_identifier"]
    )
    for symbol, source_identifier, spot, _drift, _country in PAIR_SPECS:
        scale = max(abs(spot) * 0.00004, 0.00005)
        for days_ago in range(price_days):
            for offset, adjustment in (
                (60, -2.0),
                (30, -1.2),
                (10, -0.4),
                (5, -0.15),
                (2, 0.0),
            ):
                mid = spot - days_ago * _drift + adjustment * scale
                prices.append(
                    [
                        _excel_time(as_of - timedelta(days=days_ago, minutes=offset)),
                        mid,
                        mid - scale,
                        mid + scale,
                        mid,
                        "SYNTHETIC_TEST",
                        source_identifier,
                    ]
                )

    bars = workbook.create_sheet("market_bars")
    bars.append(
        [
            "bar_time",
            "frequency",
            "open",
            "high",
            "low",
            "close",
            "source",
            "source_identifier",
        ]
    )
    for pair_index, (_symbol, source_identifier, spot, drift, _country) in enumerate(
        PAIR_SPECS
    ):
        scale = max(abs(spot) * 0.0012, 0.0008)
        for index in range(daily_bars, 0, -1):
            close = (
                spot
                - 0.065 * scale
                + index * drift
                + math.sin((index + pair_index) / 4) * scale
            )
            previous = close - drift * 0.8
            bars.append(
                [
                    _excel_time(as_of - timedelta(days=index)),
                    "daily",
                    previous,
                    close + scale * 0.7,
                    previous - scale * 0.6,
                    close,
                    "SYNTHETIC_TEST",
                    source_identifier,
                ]
            )
        for index in range(hourly_bars, 0, -1):
            close = (
                spot
                - 0.02 * scale
                + (hourly_bars - index) * drift / 24
                + math.sin((index + pair_index) / 7) * scale * 0.45
            )
            previous = close - drift / 24
            bars.append(
                [
                    _excel_time(as_of - timedelta(hours=index)),
                    "hourly",
                    previous,
                    close + scale * 0.22,
                    previous - scale * 0.18,
                    close,
                    "SYNTHETIC_TEST",
                    source_identifier,
                ]
            )

    macro = workbook.create_sheet("macro_observations")
    macro.append(
        [
            "metric_id",
            "release_time",
            "frequency",
            "value",
            "forecast_value",
            "previous_value",
            "revised_value",
            "unit",
            "source",
            "country",
        ]
    )
    for country, (policy, gdp, unemployment, cpi) in COUNTRY_PARAMS.items():
        # Keep metric IDs compatible with source.instrument_metric_link. The
        # evidence factory also accepts the suffixes, so the same workbook can
        # be used for EURUSD and the other pairs in the legacy test setup.
        metrics = (
            ("INTEREST_RATE", policy, "percent", 0.05),
            (
                "PMI_MANUFACTURING",
                53.0 if country == "US" else 50.8,
                "index",
                0.4,
            ),
            (
                "PMI_SERVICES",
                54.0 if country == "US" else 51.2,
                "index",
                0.35,
            ),
            ("UNEMPLOYMENT", unemployment, "percent", 0.12),
            ("CPI_YOY", cpi, "percent", 0.08),
            ("CORE_CPI_YOY", max(cpi - 0.35, 0.1), "percent", 0.06),
        )
        for offset, (metric, value, unit, surprise) in enumerate(metrics):
            for vintage in range(macro_vintages):
                release = as_of - timedelta(days=2 + offset + vintage * 30)
                vintage_value = value - vintage * surprise * 0.35
                macro.append(
                    [
                        f"{country}_{metric}",
                        _excel_time(release),
                        "monthly",
                        vintage_value,
                        vintage_value - surprise,
                        vintage_value - surprise * 2,
                        None,
                        unit,
                        "SYNTHETIC_TEST",
                        country,
                    ]
                )

    news = workbook.create_sheet("news_articles")
    news.append(["article_id", "publish_time", "title", "source", "related_entities"])
    article_index = 1
    for symbol, _source_identifier, _spot, _drift, country in PAIR_SPECS:
        base = symbol[:3]
        quote = symbol[3:]
        article_templates = (
            "central bank outlook shapes relative-rate view",
            "data releases test the current market trend",
            "extends its latest move as risk sentiment shifts",
            "traders monitor upcoming macro catalysts",
            "options market prices a change in volatility",
            "investors reassess growth expectations",
            "regional inflation data enters the policy debate",
            "cross-asset flows influence the session bias",
            "technical breakout attempt draws fresh attention",
            "economic survey points to a mixed outlook",
            "liquidity conditions remain in focus",
            "strategists update the medium-term scenario",
        )
        for index, template in enumerate(
            (article_templates * ((news_per_pair + len(article_templates) - 1) // len(article_templates)))[:news_per_pair],
            start=1,
        ):
            hours = index * 48
            title = f"{base}/{quote} {template}"
            news.append(
                [
                    f"SYNTH-N{article_index}",
                    _excel_time(as_of - timedelta(hours=hours)),
                    title,
                    "SYNTHETIC_TEST",
                    json.dumps(
                        {"query_tag": country, "pair": symbol},
                        ensure_ascii=False,
                    ),
                ]
            )
            article_index += 1

    workbook.save(output)
    return output


def _excel_time(value: datetime) -> datetime:
    """Excel stores datetimes without timezone metadata."""
    return value.replace(tzinfo=None)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("agent/outputs/fx-debate-synthetic/complete_multi_pair.xlsx"),
    )
    parser.add_argument("--as-of", default=DEFAULT_AS_OF)
    parser.add_argument("--price-days", type=int, default=DEFAULT_PRICE_DAYS)
    parser.add_argument("--daily-bars", type=int, default=DEFAULT_DAILY_BARS)
    parser.add_argument("--hourly-bars", type=int, default=DEFAULT_HOURLY_BARS)
    parser.add_argument("--macro-vintages", type=int, default=DEFAULT_MACRO_VINTAGES)
    parser.add_argument("--news-per-pair", type=int, default=DEFAULT_NEWS_PER_PAIR)
    args = parser.parse_args()
    as_of = datetime.fromisoformat(args.as_of.replace("Z", "+00:00"))
    if as_of.tzinfo is None:
        parser.error("--as-of must include timezone")
    path = build_workbook(
        args.output.expanduser().resolve(),
        as_of.astimezone(timezone.utc),
        price_days=args.price_days,
        daily_bars=args.daily_bars,
        hourly_bars=args.hourly_bars,
        macro_vintages=args.macro_vintages,
        news_per_pair=args.news_per_pair,
    )
    print(
        json.dumps(
            {"path": str(path), "synthetic": True, "as_of": as_of.isoformat()},
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
